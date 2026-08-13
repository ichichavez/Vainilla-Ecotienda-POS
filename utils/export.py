from __future__ import annotations
import csv
from tkinter import filedialog

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def save_dialog(default_name: str, ext: str) -> str | None:
    """Open a save-as dialog. Returns filepath or None if cancelled."""
    filetypes = {
        "csv":  [("CSV files", "*.csv"), ("All files", "*.*")],
        "xlsx": [("Excel files", "*.xlsx"), ("All files", "*.*")],
    }.get(ext, [("All files", "*.*")])

    path = filedialog.asksaveasfilename(
        defaultextension=f".{ext}",
        filetypes=filetypes,
        initialfile=default_name,
    )
    return path if path else None


def export_csv(filepath: str, headers: list, rows: list[list]) -> None:
    """Write headers + rows to a CSV file (UTF-8 with BOM for Excel compatibility)."""
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)


def export_excel(filepath: str, sheets: list[dict]) -> None:
    """
    Write one or more sheets to an .xlsx file.

    sheets: list of {
        "title":   str,
        "headers": list[str],
        "rows":    list[list],
    }
    """
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl no está instalado. Ejecutá: pip install openpyxl")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    header_fill = PatternFill("solid", fgColor="2E4057")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill    = PatternFill("solid", fgColor="F0F4F8")

    for sheet_def in sheets:
        ws = wb.create_sheet(title=sheet_def["title"][:31])
        headers = sheet_def["headers"]
        rows    = sheet_def["rows"]

        # Header row
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        for row_idx, row in enumerate(rows, start=2):
            fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if fill:
                    cell.fill = fill
                cell.alignment = Alignment(vertical="center")

        # Auto-fit column widths (approximate)
        for col_idx, header in enumerate(headers, start=1):
            max_len = len(str(header))
            for row in rows:
                if col_idx - 1 < len(row):
                    max_len = max(max_len, len(str(row[col_idx - 1] or "")))
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = min(max_len + 4, 50)

        ws.freeze_panes = "A2"

    wb.save(filepath)
